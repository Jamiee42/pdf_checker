# PDF Checker 第 11 版 GUI版（全局滚动）
import customtkinter as ctk
from tkinter import filedialog, messagebox
import pdfplumber
from PyPDF2 import PdfReader
import re
import time
import os
import sys
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import threading

# 设置外观模式
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

# ----------------------------
# 配置参数
# ----------------------------
THRESHOLD = 50
MIN_BODY_TEXT_LENGTH = 20
CONFIDENCE_MODE = True

# ----------------------------
# PDF处理核心类
# ----------------------------
class PDFCheckerCore:
    @staticmethod
    def check_pdf_status(file_path):
        try:
            reader = PdfReader(file_path)
            if reader.is_encrypted:
                return "encrypted", "PDF已加密，无法打开"
            return "ok", None
        except Exception as e:
            error_msg = str(e)
            if "PyCryptodome" in error_msg or "AES" in error_msg:
                return "encrypted", "PDF已加密，无法打开"
            return "error", f"{error_msg[:50]}"
    
    @staticmethod
    def is_cid_text(text):
        if not text:
            return False
        cleaned = text.strip()
        if not cleaned:
            return False
        cid_pattern = re.compile(r'\(cid:\d+\)|cid:\d+')
        cid_matches = cid_pattern.findall(cleaned)
        if len(cid_matches) == 0:
            return False
        total_cid_chars = sum(len(match) for match in cid_matches)
        total_chars = len(cleaned)
        if total_chars > 0:
            if len(cid_matches) > total_chars * 0.2 or total_cid_chars / total_chars > 0.4:
                return True
        return False
    
    @staticmethod
    def has_readable_text(text):
        if not text:
            return False
        cid_pattern = re.compile(r'\(cid:\d+\)|cid:\d+')
        cleaned = cid_pattern.sub('', text)
        cleaned = re.sub(r'\s+', '', cleaned)
        if not cleaned:
            return False
        printable_count = 0
        for char in cleaned:
            if ord(char) >= 32:
                printable_count += 1
        return printable_count >= 1
    
    @staticmethod
    def is_curved_page_by_font(page_num, pdf_reader):
        try:
            page = pdf_reader.pages[page_num - 1]
            if '/Resources' in page:
                resources = page['/Resources']
                if '/Font' in resources:
                    fonts = resources['/Font']
                    has_editable_font = False
                    for font_key in fonts.keys():
                        font_obj = fonts[font_key]
                        if hasattr(font_obj, 'get_object'):
                            font_obj = font_obj.get_object()
                        if '/Subtype' in font_obj:
                            font_type = font_obj['/Subtype']
                            if font_type not in ['/Type3']:
                                has_editable_font = True
                                break
                    if not has_editable_font:
                        return True
                    return False
                else:
                    return True
            else:
                return True
        except Exception:
            return False
    
    @staticmethod
    def classify_page(page, page_num, pdf_reader, confidence_mode=True):
        text = page.extract_text()
        has_images = len(page.images) > 0
        vector_count = len(page.rects) + len(page.lines) + len(page.curves)
        is_cid = PDFCheckerCore.is_cid_text(text) if text else False
        has_readable = PDFCheckerCore.has_readable_text(text) if text else False
        is_curved = False
        if confidence_mode:
            is_curved = PDFCheckerCore.is_curved_page_by_font(page_num, pdf_reader)
        
        if not text or not text.strip():
            if has_images:
                return "不可编辑", ""
            return "空白页", ""
        
        if is_cid:
            if confidence_mode:
                return "不可编辑", "疑似转曲（CID格式）"
            return "不可编辑", ""
        
        if not has_readable:
            if confidence_mode:
                return "不可编辑", "无可识别字符，疑似转曲或乱码"
            return "不可编辑", ""
        
        if is_curved:
            if confidence_mode:
                return "不可编辑", "疑似转曲（字体异常）"
            return "不可编辑", ""
        
        if len(text.strip()) < MIN_BODY_TEXT_LENGTH and not has_images and vector_count < 10:
            if confidence_mode:
                return "空白页", "仅含少量文字（疑似页眉页脚）"
            return "空白页", ""
        
        if len(text.strip()) >= THRESHOLD:
            return "可编辑", ""
        
        if has_images and len(text.strip()) < THRESHOLD:
            if confidence_mode:
                return "不可编辑", "含图像且文本较少"
            return "不可编辑", ""
        
        if len(text.strip()) >= 10:
            return "可编辑", ""
        
        return "不可编辑", ""
    
    @staticmethod
    def analyze(pdf_path, progress_callback=None):
        status, error_msg = PDFCheckerCore.check_pdf_status(pdf_path)
        
        if status != "ok":
            return {
                "success": False,
                "error": error_msg,
                "filename": os.path.basename(pdf_path)
            }
        
        with pdfplumber.open(pdf_path) as pdf:
            total_pages = len(pdf.pages)
            pdf_reader = PdfReader(pdf_path)
            
            stats = {
                "可编辑": {"pages": [], "count": 0},
                "不可编辑": {"pages": [], "count": 0},
                "空白页": {"pages": [], "count": 0},
                "无法识别": {"pages": [], "count": 0}
            }
            
            review_items = []
            
            for i, page in enumerate(pdf.pages, start=1):
                page_type, remark = PDFCheckerCore.classify_page(page, i, pdf_reader, CONFIDENCE_MODE)
                
                if remark:
                    review_items.append((i, remark))
                
                stats[page_type]["pages"].append(i)
                stats[page_type]["count"] += 1
                
                if progress_callback:
                    progress_callback(i, total_pages)
            
            notice_parts = []
            if review_items:
                review_by_type = {}
                for page_num, review_type in review_items:
                    if review_type not in review_by_type:
                        review_by_type[review_type] = []
                    review_by_type[review_type].append(page_num)
                
                for review_type, pages in review_by_type.items():
                    pages_str = ",".join(map(str, pages))
                    notice_parts.append(f"建议复核：{review_type}（第{pages_str}页）")
            
            remark_text = "；".join(notice_parts) if notice_parts else ""
            
            return {
                "success": True,
                "filename": os.path.basename(pdf_path),
                "total_pages": total_pages,
                "stats": stats,
                "remark": remark_text
            }
    
    @staticmethod
    def export_to_excel(result, output_path):
        wb = Workbook()
        ws = wb.active
        ws.title = "PDF 页面统计"
        
        ws.merge_cells("A1:A2")
        ws.merge_cells("B1:B2")
        ws.merge_cells("C1:E1")
        ws.merge_cells("F1:H1")
        ws.merge_cells("I1:K1")
        ws.merge_cells("L1:N1")
        ws.merge_cells("O1:O2")
        
        ws["A1"] = "文件名"
        ws["B1"] = "总页数"
        ws["C1"] = "可编辑"
        ws["F1"] = "不可编辑"
        ws["I1"] = "空白页"
        ws["L1"] = "无法识别"
        ws["O1"] = "注意"
        
        page_types = ["可编辑", "不可编辑", "空白页", "无法识别"]
        start_col = 3
        for pt in page_types:
            ws.cell(row=2, column=start_col, value="页数")
            ws.cell(row=2, column=start_col + 1, value="页码")
            ws.cell(row=2, column=start_col + 2, value="占比")
            start_col += 3
        
        fill_green = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        for row in ws["A1:O2"]:
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(bold=True)
                cell.fill = fill_green
        
        stats = result["stats"]
        total = result["total_pages"]
        
        ws.append([
            result["filename"],
            total,
            stats["可编辑"]["count"],
            ",".join(map(str, stats["可编辑"]["pages"])),
            f"{stats['可编辑']['count']/total:.1%}" if total > 0 else "0%",
            stats["不可编辑"]["count"],
            ",".join(map(str, stats["不可编辑"]["pages"])),
            f"{stats['不可编辑']['count']/total:.1%}" if total > 0 else "0%",
            stats["空白页"]["count"],
            ",".join(map(str, stats["空白页"]["pages"])),
            f"{stats['空白页']['count']/total:.1%}" if total > 0 else "0%",
            stats["无法识别"]["count"],
            ",".join(map(str, stats["无法识别"]["pages"])),
            f"{stats['无法识别']['count']/total:.1%}" if total > 0 else "0%",
            result["remark"]
        ])
        
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for row in ws.iter_rows(min_row=1, max_col=15, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
        
        hidden_columns = ['D', 'E', 'G', 'H', 'J', 'K', 'M', 'N']
        for col in hidden_columns:
            ws.column_dimensions[col].hidden = True
        
        wb.save(output_path)
        return output_path


# ----------------------------
# GUI主窗口（全局滚动）
# ----------------------------
class PDFCheckerApp:
    def __init__(self):
        self.window = ctk.CTk()
        self.window.title("PDF Checker 预检助手")
        self.window.geometry("950x750")
        self.window.minsize(800, 600)

        self.set_window_icon()
        
        self.pdf_path = None
        self.output_path = None
        self.analysis_result = None
        
        self.setup_ui()
    
    def set_window_icon(self):
        """设置窗口图标（标题栏和任务栏）"""
        try:
            # 获取图标文件的路径
            if getattr(sys, 'frozen', False):
                # 打包成exe后，图标在临时目录
                base_path = sys._MEIPASS
            else:
                # 开发环境，图标在当前目录
                base_path = os.path.dirname(os.path.abspath(__file__))
            
            icon_path = os.path.join(base_path, "icon.ico")
            
            if os.path.exists(icon_path):
                self.window.iconbitmap(icon_path)
        except Exception:
            pass

    def setup_ui(self):
        # 创建全局滚动框架
        self.main_scrollable = ctk.CTkScrollableFrame(self.window, label_text="")
        self.main_scrollable.pack(fill="both", expand=True, padx=10, pady=10)
        
        # 标题
        title_label = ctk.CTkLabel(
            self.main_scrollable, 
            text="PDF Checker 预检助手", 
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(pady=(10, 5))

        # 版本号
        self.window.title("PDF Checker 预检助手  v1.0.0")
        
        subtitle_label = ctk.CTkLabel(
            self.main_scrollable,
            text="PDF页面类型分析工具 - 可编辑/不可编辑/空白页统计",
            font=ctk.CTkFont(size=12)
        )
        subtitle_label.pack(pady=(0, 15))
        
        # 分隔线
        separator = ctk.CTkFrame(self.main_scrollable, height=2, fg_color="gray70")
        separator.pack(fill="x", padx=10, pady=5)
        
        # 文件选择区域
        file_frame = ctk.CTkFrame(self.main_scrollable)
        file_frame.pack(pady=10, padx=10, fill="x")
        
        ctk.CTkLabel(file_frame, text="📄 1. 选择PDF文件", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", pady=5)
        
        file_select_frame = ctk.CTkFrame(file_frame)
        file_select_frame.pack(fill="x", pady=5)
        
        self.file_path_var = ctk.StringVar(value="未选择文件")
        self.file_label = ctk.CTkLabel(file_select_frame, textvariable=self.file_path_var, fg_color="gray90", corner_radius=5, height=35)
        self.file_label.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        self.select_btn = ctk.CTkButton(file_select_frame, text="浏览", width=80, command=self.select_pdf)
        self.select_btn.pack(side="right")
        
        # 输出路径选择区域
        output_frame = ctk.CTkFrame(self.main_scrollable)
        output_frame.pack(pady=10, padx=10, fill="x")
        
        ctk.CTkLabel(output_frame, text="💾 2. 选择输出位置", font=ctk.CTkFont(size=14, weight="bold")).pack(anchor="w", pady=5)
        
        output_select_frame = ctk.CTkFrame(output_frame)
        output_select_frame.pack(fill="x", pady=5)
        
        self.output_path_var = ctk.StringVar(value="未选择输出位置")
        self.output_label = ctk.CTkLabel(output_select_frame, textvariable=self.output_path_var, fg_color="gray90", corner_radius=5, height=35)
        self.output_label.pack(side="left", fill="x", expand=True, padx=(0, 10))
        
        self.output_btn = ctk.CTkButton(output_select_frame, text="浏览", width=80, command=self.select_output)
        self.output_btn.pack(side="right")
        
        # 开始按钮
        self.start_btn = ctk.CTkButton(
            self.main_scrollable, 
            text="▶ 开始分析", 
            font=ctk.CTkFont(size=16, weight="bold"),
            height=45,
            command=self.start_analysis
        )
        self.start_btn.pack(pady=15, padx=10, fill="x")
        
        # 进度条区域
        progress_frame = ctk.CTkFrame(self.main_scrollable)
        progress_frame.pack(pady=5, padx=10, fill="x")
        
        self.progress_bar = ctk.CTkProgressBar(progress_frame)
        self.progress_bar.pack(fill="x", pady=5)
        self.progress_bar.set(0)
        
        self.progress_label = ctk.CTkLabel(progress_frame, text="")
        self.progress_label.pack()
        
        # 分隔线
        separator2 = ctk.CTkFrame(self.main_scrollable, height=2, fg_color="gray70")
        separator2.pack(fill="x", padx=10, pady=10)
        
        # 结果展示区域
        result_label = ctk.CTkLabel(self.main_scrollable, text="📊 3. 分析结果", font=ctk.CTkFont(size=14, weight="bold"))
        result_label.pack(anchor="w", padx=10, pady=(0, 5))
        
        # 统计卡片区域
        stats_frame = ctk.CTkFrame(self.main_scrollable)
        stats_frame.pack(fill="x", pady=5, padx=10)
        
        self.stats_cards = {}
        colors = {"可编辑": "#2ecc71", "不可编辑": "#e74c3c", "空白页": "#95a5a6", "无法识别": "#f39c12"}
        
        # 使用grid布局，4列
        for i, (pt, color) in enumerate(colors.items()):
            card = ctk.CTkFrame(stats_frame, fg_color=color, corner_radius=10)
            card.grid(row=0, column=i, padx=5, pady=5, sticky="nsew")
            stats_frame.grid_columnconfigure(i, weight=1)
            
            count_label = ctk.CTkLabel(card, text="0", font=ctk.CTkFont(size=28, weight="bold"), text_color="white")
            count_label.pack(pady=(10, 5))
            
            type_label = ctk.CTkLabel(card, text=pt, font=ctk.CTkFont(size=12), text_color="white")
            type_label.pack()
            
            percent_label = ctk.CTkLabel(card, text="0%", font=ctk.CTkFont(size=12), text_color="white")
            percent_label.pack(pady=(0, 10))
            
            self.stats_cards[pt] = {"count": count_label, "percent": percent_label}
        
        # 详细信息区域
        detail_label = ctk.CTkLabel(self.main_scrollable, text="详细信息", font=ctk.CTkFont(size=12, weight="bold"))
        detail_label.pack(anchor="w", padx=10, pady=(10, 5))
        
        self.detail_text = ctk.CTkTextbox(self.main_scrollable, height=200)
        self.detail_text.pack(fill="x", padx=10, pady=5)
        
        # 导出按钮
        self.export_btn = ctk.CTkButton(
            self.main_scrollable,
            text="📎 导出Excel",
            height=40,
            command=self.export_excel,
            state="disabled"
        )
        self.export_btn.pack(pady=15, padx=10, fill="x")
        
        # 底部留白
        ctk.CTkLabel(self.main_scrollable, text="").pack(pady=10)
    
    def select_pdf(self):
        file_path = filedialog.askopenfilename(
            title="选择PDF文件",
            filetypes=[("PDF文件", "*.pdf"), ("所有文件", "*.*")]
        )
        if file_path:
            self.pdf_path = file_path
            self.file_path_var.set(os.path.basename(file_path))
    
    def select_output(self):
        if not self.pdf_path:
            messagebox.showwarning("提示", "请先选择PDF文件")
            return
        
        base_name = os.path.splitext(os.path.basename(self.pdf_path))[0]
        default_name = f"{base_name}_Result.xlsx"
        
        output_path = filedialog.asksaveasfilename(
            title="保存Excel文件",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile=default_name
        )
        if output_path:
            self.output_path = output_path
            self.output_path_var.set(os.path.basename(output_path))
    
    def update_progress(self, current, total):
        progress = current / total
        self.progress_bar.set(progress)
        self.progress_label.configure(text=f"处理中: {current}/{total} 页 ({progress:.1%})")
        self.window.update_idletasks()
    
    def start_analysis(self):
        if not self.pdf_path:
            messagebox.showwarning("提示", "请先选择PDF文件")
            return
        
        if not self.output_path:
            messagebox.showwarning("提示", "请先选择输出位置")
            return
        
        for pt in self.stats_cards:
            self.stats_cards[pt]["count"].configure(text="0")
            self.stats_cards[pt]["percent"].configure(text="0%")
        self.detail_text.delete("1.0", "end")
        self.export_btn.configure(state="disabled")
        
        self.start_btn.configure(state="disabled", text="分析中...")
        self.progress_bar.set(0)
        
        thread = threading.Thread(target=self.run_analysis, daemon=True)
        thread.start()
    
    def run_analysis(self):
        try:
            result = PDFCheckerCore.analyze(self.pdf_path, self.update_progress)
            
            if not result["success"]:
                self.window.after(0, lambda: messagebox.showerror("错误", result["error"]))
                return
            
            self.analysis_result = result
            self.window.after(0, self.update_results)
            
        except Exception as e:
            self.window.after(0, lambda: messagebox.showerror("错误", f"分析失败：{str(e)}"))
        finally:
            self.window.after(0, self.finish_analysis)
    
    def update_results(self):
        result = self.analysis_result
        stats = result["stats"]
        total = result["total_pages"]
        
        for pt in ["可编辑", "不可编辑", "空白页", "无法识别"]:
            count = stats[pt]["count"]
            percent = f"{count/total:.1%}" if total > 0 else "0%"
            self.stats_cards[pt]["count"].configure(text=str(count))
            self.stats_cards[pt]["percent"].configure(text=percent)
        
        detail = f"文件名：{result['filename']}\n"
        detail += f"总页数：{total}\n"
        detail += f"{'='*50}\n"
        detail += f"可编辑：{stats['可编辑']['count']} 页（页码：{stats['可编辑']['pages']}）\n"
        detail += f"不可编辑：{stats['不可编辑']['count']} 页（页码：{stats['不可编辑']['pages']}）\n"
        detail += f"空白页：{stats['空白页']['count']} 页（页码：{stats['空白页']['pages']}）\n"
        detail += f"无法识别：{stats['无法识别']['count']} 页（页码：{stats['无法识别']['pages']}）\n"
        if result["remark"]:
            detail += f"{'='*50}\n"
            detail += f"注意：{result['remark']}\n"
        
        self.detail_text.delete("1.0", "end")
        self.detail_text.insert("1.0", detail)
        self.export_btn.configure(state="normal")
    
    def finish_analysis(self):
        self.start_btn.configure(state="normal", text="▶ 开始分析")
        self.progress_label.configure(text="分析完成")
    
    def export_excel(self):
        if not self.analysis_result:
            return
        
        try:
            PDFCheckerCore.export_to_excel(self.analysis_result, self.output_path)
            messagebox.showinfo("成功", f"Excel已导出至：\n{self.output_path}")
        except Exception as e:
            messagebox.showerror("错误", f"导出失败：{str(e)}")
    
    def run(self):
        self.window.mainloop()


if __name__ == "__main__":
    app = PDFCheckerApp()
    app.run()